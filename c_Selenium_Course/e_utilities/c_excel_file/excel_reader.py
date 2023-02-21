import openpyxl

workbook = openpyxl.load_workbook("../testData/Test_Data.xlsx")
sheet = workbook.active
print(sheet['A2'].value)
print("---------------------------------------------")
print(str(int(sheet.cell(row=3, column=3).value)))
print("---------------------------------------------")
print((sheet["A1:E5"]))
print("---------------------------------------------")
# Access all cells from A1 to D5
for row in sheet["A1:E5"]:
    print([x.value for x in row])
print("---------------------------------------------")
# Access all cells in column A
for data in sheet["A"]:
    if str(data.value) == 'None':
        break
    else:
        print(data.value)
print("----------------------------------------------")
print(f"Max row in the active sheet is {sheet.max_row}")
print(f"Max column in the active sheet is {sheet.max_column}")
print("----------------------------------------------")
# Access all cells from between
# "Row 1 and Row 2" and "Column 1 and Column 3"
for row in sheet.iter_rows(min_row=2,
                           max_row=3,
                           min_col=1,
                           max_col=4):
    print([data.value for data in row])

